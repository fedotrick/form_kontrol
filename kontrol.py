import sys
import pandas as pd
from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QFormLayout, QLineEdit,
    QDateEdit, QPushButton, QMessageBox, QGroupBox, QLabel, QScrollArea, QComboBox, QHBoxLayout, QGraphicsDropShadowEffect
)
from PySide6.QtCore import QDate, Qt, QPropertyAnimation, QEasingCurve, QEvent
from PySide6 import QtGui
from PySide6.QtGui import QFont, QColor
import os
from openpyxl import load_workbook, Workbook
from datetime import datetime, timedelta

class ControlForm(QWidget):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Электронный журнал контроля")
        self.setGeometry(100, 100, 900, 950)

        layout = QVBoxLayout()
        layout.setSpacing(5)
        layout.setContentsMargins(5, 5, 5, 5)

        # Добавляем общие стили в стиле Dracula
        self.setStyleSheet("""
            QWidget {
                background-color: #282a36;
                color: #f8f8f2;
                font-family: 'Segoe UI', 'Aptos';
                font-size: 11px;
            }
            
            QGroupBox {
                border: 1px solid #44475a;
                border-radius: 4px;
                margin-top: 0.5em;
                padding: 8px;
            }
            
            QLineEdit, QComboBox, QDateEdit {
                background-color: #44475a;
                color: #f8f8f2;
                border: 1px solid #6272a4;
                border-radius: 2px;
                padding: 3px;
                min-width: 80px;
                max-width: 120px;
                height: 20px;
            }
            
            QLineEdit:focus, QComboBox:focus, QDateEdit:focus {
                border: 2px solid #bd93f9;
            }
            
            QLabel {
                color: #f8f8f2;
                padding: 2px;
            }
            
            QPushButton {
                background-color: #6272a4;
                color: #f8f8f2;
                border: none;
                padding: 5px 10px;
                border-radius: 2px;
                height: 25px;
            }
            
            QPushButton:hover {
                background-color: #bd93f9;
            }
            
            QPushButton:pressed {
                background-color: #ff79c6;
            }
            
            QScrollArea {
                border: none;
            }
            
            QScrollBar:vertical {
                background-color: #282a36;
                width: 14px;
                margin: 15px 0;
            }
            
            QScrollBar::handle:vertical {
                background-color: #44475a;
                min-height: 30px;
                border-radius: 7px;
            }
            
            QScrollBar::handle:vertical:hover {
                background-color: #6272a4;
            }
            
            QComboBox::drop-down {
                border: none;
                width: 20px;
            }
            
            QComboBox QAbstractItemView {
                background-color: #44475a;
                color: #f8f8f2;
                selection-background-color: #6272a4;
            }
        """)

        # Создание области прокрутки
        scroll_area = QScrollArea(self)
        scroll_area.setWidgetResizable(True)
        
        # Основной виджет для прокрутки
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)

        # Блок 1: Основные данные
        group_box1 = QGroupBox("Основные данные")
        group_box1.setStyleSheet("""
            QGroupBox::title {
                color: #bd93f9;
            }
        """)
        form_layout1 = QFormLayout()
        form_layout1.setSpacing(3)
        form_layout1.setContentsMargins(5, 5, 5, 5)
        
        # Выпадающий список для номера плавки
        self.номер_плавки_input = QComboBox(self)
        self.load_plavka_numbers()
        
        # Добавляем поле для отображения наименования отливки (только для чтения)
        self.наименование_отливки_input = QLineEdit(self)
        self.наименование_отливки_input.setReadOnly(True)  # Только для чтения
        self.наименование_отливки_input.setStyleSheet("""
            QLineEdit {
                background-color: #44475a;
                color: #f8f8f2;
                border: 1px solid #6272a4;
                border-radius: 2px;
                padding: 3px;
                min-width: 200px;
                max-width: 300px;
                height: 20px;
            }
        """)
        
        # Список участников
        persons = [
            "Елхова", "Лабуткина", "Рябова", "Улитина"            
        ]
        # Сортировка списка участников по возрастанию
        persons.sort()
        
        # Создаем основные поля ввода
        self.контроль_отлито_input = QLineEdit(self)
        self.контроль_принято_input = QLineEdit(self)
        self.контроль_принято_input.setReadOnly(True)  # Только для чтения
        
        # Создаем комбобоксы для контролеров
        self.контролер1_input = QComboBox(self)
        self.контролер1_input.addItems(persons)  # Добавляем участников в комбобокс
        self.контролер1_input.setFont(QtGui.QFont("Aptos", 12, QtGui.QFont.Bold))
        self.контролер1_input.setStyleSheet("color: white;")
        self.контролер1_input.setCurrentIndex(-1)  # Ничего не выбрано по умолчанию
        
        self.контролер2_input = QComboBox(self)
        self.контролер2_input.addItems(persons)  # Добавляем участников в комбобокс
        self.контролер2_input.setFont(QtGui.QFont("Aptos", 12, QtGui.QFont.Bold))
        self.контролер2_input.setStyleSheet("color: white;")
        self.контролер2_input.setCurrentIndex(-1)

        self.контролер3_input = QComboBox(self)
        self.контролер3_input.addItems(persons)  # Добавляем участников в комбобокс
        self.контролер3_input.setFont(QtGui.QFont("Aptos", 12, QtGui.QFont.Bold))
        self.контролер3_input.setStyleSheet("color: white;")
        self.контролер3_input.setCurrentIndex(-1)
        
        # Настраиваем поле даты
        self.контроль_дата_приемки_input = QDateEdit(self)
        self.контроль_дата_приемки_input.setDisplayFormat("dd.MM.yyyy")
        self.контроль_дата_приемки_input.setDate(QDate.currentDate())  # Текущая дата по умолчанию
        self.контроль_дата_приемки_input.setCalendarPopup(True)  # Разрешаем всплывающий календарь
        self.контроль_дата_приемки_input.setStyleSheet("""
            QDateEdit {
                padding: 8px;
                font-size: 12px;
                min-width: 120px;
                border-radius: 4px;
                background-color: #44475a;
                color: #f8f8f2;
            }
        """)
        
        # Добавляем остальные поля в форму
        form_layout1.addRow(QLabel("Дата приемки:"), self.контроль_дата_приемки_input)
        form_layout1.addRow(QLabel("Номер плавки:"), self.номер_плавки_input)
        form_layout1.addRow(QLabel("Наименование отливки:"), self.наименование_отливки_input)
        form_layout1.addRow(QLabel("Отлито, шт.:"), self.контроль_отлито_input)
        form_layout1.addRow(QLabel("Принято, шт.:"), self.контроль_принято_input)
        form_layout1.addRow(QLabel("Контролер 1:"), self.контролер1_input)
        form_layout1.addRow(QLabel("Контролер 2:"), self.контролер2_input)
        form_layout1.addRow(QLabel("Контролер 3:"), self.контролер3_input)
        
        group_box1.setLayout(form_layout1)

        # Блок 2: Второй сорт
        group_box2 = QGroupBox("Второй сорт")
        group_box2.setStyleSheet("""
            QGroupBox::title {
                color: #50fa7b;
            }
        """)
        form_layout2 = QFormLayout()
        form_layout2.setSpacing(3)
        form_layout2.setContentsMargins(5, 5, 5, 5)
        
        self.второй_сорт_раковины_input = QLineEdit(self)
        self.второй_сорт_зарез_input = QLineEdit(self)

        form_layout2.addRow(QLabel("Раковины:"), self.второй_сорт_раковины_input)
        form_layout2.addRow(QLabel("Зарез:"), self.второй_сорт_зарез_input)

        group_box2.setLayout(form_layout2)

        # Блок 3: Доработка
        group_box3 = QGroupBox("Доработка")
        group_box3.setStyleSheet("""
            QGroupBox::title {
                color: #ffb86c;
            }
        """)
        form_layout3 = QFormLayout()
        form_layout3.setSpacing(3)
        form_layout3.setContentsMargins(5, 5, 5, 5)
        
        self.доработка_раковины_input = QLineEdit(self)
        self.доработка_раковины_input.hide()  # Hide the field
        self.доработка_зарез_input = QLineEdit(self)
        self.доработка_зарез_input.hide()  # Hide the field
        self.доработка_несоответствие_размеров_input = QLineEdit(self)
        self.доработка_несоответствие_внешнего_вида_input = QLineEdit(self)
        self.доработка_наплыв_металла_input = QLineEdit(self)
        self.доработка_прорыв_металла_input = QLineEdit(self)
        self.доработка_вырыв_input = QLineEdit(self)
        self.доработка_облой_input = QLineEdit(self)
        self.доработка_песок_на_поверхности_input = QLineEdit(self)
        self.доработка_песок_в_резьбе_input = QLineEdit(self)
        self.доработка_клей_input = QLineEdit(self)
        self.доработка_коробление_input = QLineEdit(self)
        self.доработка_дефект_пеномодели_input = QLineEdit(self)
        self.доработка_лапы_input = QLineEdit(self)
        self.доработка_питатель_input = QLineEdit(self)
        self.доработка_корона_input = QLineEdit(self)
        self.доработка_смещение_input = QLineEdit(self)

        # Add rows without the hidden fields
        form_layout3.addRow(QLabel("Несоответствие размеров:"), self.доработка_несоответствие_размеров_input)
        form_layout3.addRow(QLabel("Несоответствие внешнего вида:"), self.доработка_несоответствие_внешнего_вида_input)
        form_layout3.addRow(QLabel("Наплыв металла:"), self.доработка_наплыв_металла_input)
        form_layout3.addRow(QLabel("Прорыв металла:"), self.доработка_прорыв_металла_input)
        form_layout3.addRow(QLabel("Вырыв:"), self.доработка_вырыв_input)
        form_layout3.addRow(QLabel("Облой:"), self.доработка_облой_input)
        form_layout3.addRow(QLabel("Песок на поверхности:"), self.доработка_песок_на_поверхности_input)
        form_layout3.addRow(QLabel("Песок в резьбе:"), self.доработка_песок_в_резьбе_input)
        form_layout3.addRow(QLabel("Клей:"), self.доработка_клей_input)
        form_layout3.addRow(QLabel("Коробление:"), self.доработка_коробление_input)
        form_layout3.addRow(QLabel("Дефект пеномодели:"), self.доработка_дефект_пеномодели_input)
        form_layout3.addRow(QLabel("Лапы:"), self.доработка_лапы_input)
        form_layout3.addRow(QLabel("Питатель:"), self.доработка_питатель_input)
        form_layout3.addRow(QLabel("Корона:"), self.доработка_корона_input)
        form_layout3.addRow(QLabel("Смещение:"), self.доработка_смещение_input)

        group_box3.setLayout(form_layout3)

        # Блок 4: Окончательный брак
        group_box4 = QGroupBox("Окончательный брак")
        group_box4.setStyleSheet("""
            QGroupBox::title {
                color: #ff5555;
            }
        """)
        form_layout4 = QFormLayout()
        form_layout4.setSpacing(3)
        form_layout4.setContentsMargins(5, 5, 5, 5)
        
        self.окончательный_брак_недолив_input = QLineEdit(self)
        self.окончательный_брак_раковины_input = QLineEdit(self)
        self.окончательный_брак_коробление_input = QLineEdit(self)
        self.окончательный_брак_спай_input = QLineEdit(self)
        self.окончательный_брак_трещины_input = QLineEdit(self)
        self.окончательный_брак_пригар_песка_input = QLineEdit(self)
        self.окончательный_брак_пористость_input = QLineEdit(self)
        self.окончательный_брак_вырыв_input = QLineEdit(self)
        self.окончательный_брак_скол_input = QLineEdit(self)
        self.окончательный_брак_слом_input = QLineEdit(self)
        self.окончательный_брак_зарез_input = QLineEdit(self)
        self.окончательный_брак_нарушение_геометрии_input = QLineEdit(self)
        self.окончательный_брак_рыхлота_input = QLineEdit(self)
        self.окончательный_брак_непроклей_input = QLineEdit(self)
        self.окончательный_брак_пеномодель_input = QLineEdit(self)
        self.окончательный_брак_наплыв_металла_input = QLineEdit(self)
        self.окончательный_брак_несоответствие_размеров_input = QLineEdit(self)
        self.окончательный_брак_несоответствие_внешнего_вида_input = QLineEdit(self)
        self.окончательный_брак_нарушение_маркировки_input = QLineEdit(self)
        self.окончательный_брак_неслитина_input = QLineEdit(self)
        self.окончательный_брак_прочее_input = QLineEdit(self)

        form_layout4.addRow(QLabel("Недолив:"), self.окончательный_брак_недолив_input)
        form_layout4.addRow(QLabel("Раковины:"), self.окончательный_брак_раковины_input)
        form_layout4.addRow(QLabel("Коробление:"), self.окончательный_брак_коробление_input)
        form_layout4.addRow(QLabel("Спай:"), self.окончательный_брак_спай_input)
        form_layout4.addRow(QLabel("Трещины:"), self.окончательный_брак_трещины_input)
        form_layout4.addRow(QLabel("Пригар песка:"), self.окончательный_брак_пригар_песка_input)
        form_layout4.addRow(QLabel("Пористость:"), self.окончательный_брак_пористость_input)
        form_layout4.addRow(QLabel("Вырыв:"), self.окончательный_брак_вырыв_input)
        form_layout4.addRow(QLabel("Скол:"), self.окончательный_брак_скол_input)
        form_layout4.addRow(QLabel("Слом:"), self.окончательный_брак_слом_input)
        form_layout4.addRow(QLabel("Зарез:"), self.окончательный_брак_зарез_input)
        form_layout4.addRow(QLabel("Нарушение геометрии:"), self.окончательный_брак_нарушение_геометрии_input)
        form_layout4.addRow(QLabel("Рыхлота:"), self.окончательный_брак_рыхлота_input)
        form_layout4.addRow(QLabel("Непроклей:"), self.окончательный_брак_непроклей_input)
        form_layout4.addRow(QLabel("Пеномодель:"), self.окончательный_брак_пеномодель_input)
        form_layout4.addRow(QLabel("Наплыв металла:"), self.окончательный_брак_наплыв_металла_input)
        form_layout4.addRow(QLabel("Несоответствие размеров:"), self.окончательный_брак_несоответствие_размеров_input)
        form_layout4.addRow(QLabel("Несоответствие внешнего вида:"), self.окончательный_брак_несоответствие_внешнего_вида_input)
        form_layout4.addRow(QLabel("Нарушение маркировки:"), self.окончательный_брак_нарушение_маркировки_input)
        form_layout4.addRow(QLabel("Неслитина:"), self.окончательный_брак_неслитина_input)
        form_layout4.addRow(QLabel("Прочее:"), self.окончательный_брак_прочее_input)

        group_box4.setLayout(form_layout4)

        # Создаем горизонтальные layout для группировки полей
        h_layout = QHBoxLayout()
        
        # Группируем GroupBox'ы по два в ряд
        left_column = QVBoxLayout()
        left_column.addWidget(group_box1)
        left_column.addWidget(group_box3)
        
        right_column = QVBoxLayout()
        right_column.addWidget(group_box2)
        right_column.addWidget(group_box4)
        
        h_layout.addLayout(left_column)
        h_layout.addLayout(right_column)
        
        scroll_layout.addLayout(h_layout)

        # Установка виджета прокрутки
        scroll_area.setWidget(scroll_widget)
        layout.addWidget(scroll_area)

        # Кнопка для сохранения данных
        self.save_button = QPushButton("Сохранить", self)

        # Установка параметров кнопки
        self.save_button.setStyleSheet("""
            QPushButton {
                background-color: #50fa7b;
                color: #282a36;
                font-size: 14px;
                padding: 12px 30px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #69ff94;
            }
            QPushButton:pressed {
                background-color: #41d66b;
            }
        """)

        # Подключение сигнала
        self.save_button.clicked.connect(self.save_data)

        # Добавление кнопки в layout
        layout.addWidget(self.save_button)

        self.setLayout(layout)

        # Подключение события изменения для расчета контроль_принято
        self.контроль_отлито_input.textChanged.connect(self.calculate_control_prinato)
        self.второй_сорт_раковины_input.textChanged.connect(self.calculate_control_prinato)
        self.второй_сорт_зарез_input.textChanged.connect(self.calculate_control_prinato)
        self.доработка_раковины_input.textChanged.connect(self.calculate_control_prinato)
        self.доработка_зарез_input.textChanged.connect(self.calculate_control_prinato)
        self.доработка_несоответствие_размеров_input.textChanged.connect(self.calculate_control_prinato)
        self.доработка_несоответствие_внешнего_вида_input.textChanged.connect(self.calculate_control_prinato)
        self.доработка_наплыв_металла_input.textChanged.connect(self.calculate_control_prinato)
        self.доработка_прорыв_металла_input.textChanged.connect(self.calculate_control_prinato)
        self.доработка_вырыв_input.textChanged.connect(self.calculate_control_prinato)
        self.доработка_облой_input.textChanged.connect(self.calculate_control_prinato)
        self.доработка_песок_на_поверхности_input.textChanged.connect(self.calculate_control_prinato)
        self.доработка_песок_в_резьбе_input.textChanged.connect(self.calculate_control_prinato)
        self.доработка_клей_input.textChanged.connect(self.calculate_control_prinato)
        self.доработка_коробление_input.textChanged.connect(self.calculate_control_prinato)
        self.доработка_дефект_пеномодели_input.textChanged.connect(self.calculate_control_prinato)
        self.доработка_лапы_input.textChanged.connect(self.calculate_control_prinato)
        self.доработка_питатель_input.textChanged.connect(self.calculate_control_prinato)
        self.доработка_корона_input.textChanged.connect(self.calculate_control_prinato)
        self.доработка_смещение_input.textChanged.connect(self.calculate_control_prinato)
        self.окончательный_брак_недолив_input.textChanged.connect(self.calculate_control_prinato)
        self.окончательный_брак_вырыв_input.textChanged.connect(self.calculate_control_prinato)
        self.окончательный_брак_зарез_input.textChanged.connect(self.calculate_control_prinato)
        self.окончательный_брак_коробление_input.textChanged.connect(self.calculate_control_prinato)
        self.окончательный_брак_наплыв_металла_input.textChanged.connect(self.calculate_control_prinato)
        self.окончательный_брак_нарушение_геометрии_input.textChanged.connect(self.calculate_control_prinato)
        self.окончательный_брак_нарушение_маркировки_input.textChanged.connect(self.calculate_control_prinato)
        self.окончательный_брак_непроклей_input.textChanged.connect(self.calculate_control_prinato)
        self.окончательный_брак_неслитина_input.textChanged.connect(self.calculate_control_prinato)
        self.окончательный_брак_несоответствие_внешнего_вида_input.textChanged.connect(self.calculate_control_prinato)
        self.окончательный_брак_несоответствие_размеров_input.textChanged.connect(self.calculate_control_prinato)
        self.окончательный_брак_пеномодель_input.textChanged.connect(self.calculate_control_prinato)
        self.окончательный_брак_пористость_input.textChanged.connect(self.calculate_control_prinato)
        self.окончательный_брак_пригар_песка_input.textChanged.connect(self.calculate_control_prinato)
        self.окончательный_брак_прочее_input.textChanged.connect(self.calculate_control_prinato)
        self.окончательный_брак_рыхлота_input.textChanged.connect(self.calculate_control_prinato)
        self.окончательный_брак_раковины_input.textChanged.connect(self.calculate_control_prinato)
        self.окончательный_брак_скол_input.textChanged.connect(self.calculate_control_prinato)
        self.окончательный_брак_слом_input.textChanged.connect(self.calculate_control_prinato)
        self.окончательный_брак_спай_input.textChanged.connect(self.calculate_control_prinato)
        self.окончательный_брак_трещины_input.textChanged.connect(self.calculate_control_prinato)
        
        # Добавляем анимацию при наведении на группы
        for group in [group_box1, group_box2, group_box3, group_box4]:
            group.enterEvent = lambda e, g=group: self.animate_group_hover(g, True)
            group.leaveEvent = lambda e, g=group: self.animate_group_hover(g, False)

        # Добавляем тени для групп
        for group in [group_box1, group_box2, group_box3, group_box4]:
            shadow = QGraphicsDropShadowEffect()
            shadow.setBlurRadius(15)
            shadow.setColor(QColor(0, 0, 0, 30))
            shadow.setOffset(0, 2)
            group.setGraphicsEffect(shadow)

        # Добавляем валидацию для числовых полей
        numeric_inputs = [
            self.контроль_отлито_input,
            self.второй_сорт_раковины_input,
            self.второй_сорт_зарез_input,
            self.доработка_раковины_input,
            self.доработка_зарез_input,
            self.доработка_несоответствие_размеров_input,
            self.доработка_несоответствие_внешнего_вида_input,
            self.доработка_наплыв_металла_input,
            self.доработка_прорыв_металла_input,
            self.доработка_вырыв_input,
            self.доработка_облой_input,
            self.доработка_песок_на_поверхности_input,
            self.доработка_песок_в_резьбе_input,
            self.доработка_клей_input,
            self.доработка_коробление_input,
            self.доработка_дефект_пеномодели_input,
            self.доработка_лапы_input,
            self.доработка_питатель_input,
            self.доработка_корона_input,
            self.доработка_смещение_input,
            self.окончательный_брак_недолив_input,
            self.окончательный_брак_вырыв_input,
            self.окончательный_брак_зарез_input,
            self.окончательный_брак_коробление_input,
            self.окончательный_брак_наплыв_металла_input,
            self.окончательный_брак_нарушение_геометрии_input,
            self.окончательный_брак_нарушение_маркировки_input,
            self.окончательный_брак_непроклей_input,
            self.окончательный_брак_неслитина_input,
            self.окончательный_брак_несоответствие_внешнего_вида_input,
            self.окончательный_брак_несоответствие_размеров_input,
            self.окончательный_брак_пеномодель_input,
            self.окончательный_брак_пористость_input,
            self.окончательный_брак_пригар_песка_input,
            self.окончательный_брак_прочее_input,
            self.окончательный_брак_рыхлота_input,
            self.окончательный_брак_раковины_input,
            self.окончательный_брак_скол_input,
            self.окончательный_брак_слом_input,
            self.окончательный_брак_спай_input,
            self.окончательный_брак_трещины_input
        ]
        
        for input_field in numeric_inputs:
            input_field.textChanged.connect(
                lambda text, field=input_field: field.setText(''.join(filter(str.isdigit, text)))
            )

        # Подключаем обработчик изменения номера плавки
        self.номер_плавки_input.currentTextChanged.connect(self.update_наименование_отливки)

        # Устанавливаем фокус на первое поле при запуске
        self.контроль_дата_приемки_input.setFocus()
        
        # Список всех интерактивных виджетов для навигации
        self.focusable_widgets = [
            self.контроль_дата_приемки_input,
            self.номер_плавки_input,
            self.контроль_отлито_input,
            self.контролер1_input,
            self.контролер2_input,
            self.контролер3_input,
            self.второй_сорт_раковины_input,
            self.второй_сорт_зарез_input,
            self.доработка_раковины_input,
            self.доработка_зарез_input,
            self.доработка_несоответствие_размеров_input,
            self.доработка_несоответствие_внешнего_вида_input,
            self.доработка_наплыв_металла_input,
            self.доработка_прорыв_металла_input,
            self.доработка_вырыв_input,
            self.доработка_облой_input,
            self.доработка_песок_на_поверхности_input,
            self.доработка_песок_в_резьбе_input,
            self.доработка_клей_input,
            self.доработка_коробление_input,
            self.доработка_дефект_пеномодели_input,
            self.доработка_лапы_input,
            self.доработка_питатель_input,
            self.доработка_корона_input,
            self.доработка_смещение_input,
            self.окончательный_брак_недолив_input,
            self.окончательный_брак_вырыв_input,
            self.окончательный_брак_зарез_input,
            self.окончательный_брак_коробление_input,
            self.окончательный_брак_наплыв_металла_input,
            self.окончательный_брак_нарушение_геометрии_input,
            self.окончательный_брак_нарушение_маркировки_input,
            self.окончательный_брак_непроклей_input,
            self.окончательный_брак_неслитина_input,
            self.окончательный_брак_несоответствие_внешнего_вида_input,
            self.окончательный_брак_несоответствие_размеров_input,
            self.окончательный_брак_пеномодель_input,
            self.окончательный_брак_пористость_input,
            self.окончательный_брак_пригар_песка_input,
            self.окончательный_брак_прочее_input,
            self.окончательный_брак_рыхлота_input,
            self.окончательный_брак_раковины_input,
            self.окончательный_брак_скол_input,
            self.окончательный_брак_слом_input,
            self.окончательный_брак_спай_input,
            self.окончательный_брак_трещины_input
        ]
        
        # Устанавливаем обработку клавиш для каждого виджета
        for widget in self.focusable_widgets:
            widget.installEventFilter(self)

    def eventFilter(self, obj, event):
        if event.type() == QEvent.KeyPress and obj in self.focusable_widgets:
            key = event.key()
            current_index = self.focusable_widgets.index(obj)
            
            if key == Qt.Key_Up:
                # Переход к предыдущему виджету (циклично)
                next_index = (current_index - 1) % len(self.focusable_widgets)
                self.focusable_widgets[next_index].setFocus()
                return True
                
            elif key == Qt.Key_Down:
                # Переход к следующему виджету (циклично)
                next_index = (current_index + 1) % len(self.focusable_widgets)
                self.focusable_widgets[next_index].setFocus()
                return True
                
        return super().eventFilter(obj, event)

    def load_plavka_numbers(self):
        if not os.path.exists('plavka.xlsx'):
            QMessageBox.warning(self, "Ошибка", "Файл plavka.xlsx не найден")
            return
        try:
            # Загрузка данных из plavka.xlsx
            self.df_plavka = pd.read_excel('plavka.xlsx')  # Сохраняем DataFrame как атрибут класса
            
            # Фильтрация номеров, содержащих "/25"
            self.df_plavka = self.df_plavka[self.df_plavka['Учетный_номер'].astype(str).str.contains('/25')]
            
            # Загрузка данных из control.xlsx, если файл существует
            try:
                df_control = pd.read_excel('control.xlsx')
                # Получение списка уже использованных номеров плавок
                used_numbers = df_control['Номер_плавки'].astype(str).unique()
                # Фильтрация, исключая использованные номера
                self.df_plavka = self.df_plavka[~self.df_plavka['Учетный_номер'].astype(str).isin(used_numbers)]
            except FileNotFoundError:
                pass
            
            # Добавление отфильтрованных номеров в комбобокс
            available_numbers = self.df_plavka['Учетный_номер'].astype(str).tolist()
            self.номер_плавки_input.addItems(available_numbers)
            
            QMessageBox.information(self, "Информация", 
                                  f"Доступно номеров плавок: {len(available_numbers)}")
            
        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Ошибка при загрузке номеров плавок: {str(e)}")

    def update_наименование_отливки(self, selected_number):
        """Обновляет поле наименования отливки при выборе номера плавки"""
        try:
            if selected_number and hasattr(self, 'df_plavka'):
                # Находим соответствующее наименование в DataFrame
                наименование = self.df_plavka[
                    self.df_plavka['Учетный_номер'].astype(str) == selected_number
                ]['Наименование_отливки'].iloc[0]
                self.наименование_отливки_input.setText(str(наименование))
            else:
                self.наименование_отливки_input.clear()
        except Exception as e:
            self.наименование_отливки_input.clear()
            print(f"Ошибка при обновлении наименования отливки: {str(e)}")

    def calculate_control_prinato(self):
        try:
            контроль_отлито = int(self.контроль_отлито_input.text() or 0)
            второй_сорт_раковины = int(self.второй_сорт_раковины_input.text() or 0)
            второй_сорт_зарез = int(self.второй_сорт_зарез_input.text() or 0)
            доработка_раковины = int(self.доработка_раковины_input.text() or 0)
            доработка_зарез = int(self.доработка_зарез_input.text() or 0)
            доработка_несоответствие_размеров = int(self.доработка_несоответствие_размеров_input.text() or 0)
            доработка_несоответствие_внешнего_вида = int(self.доработка_несоответствие_внешнего_вида_input.text() or 0)
            доработка_наплыв_металла = int(self.доработка_наплыв_металла_input.text() or 0)
            доработка_прорыв_металла = int(self.доработка_прорыв_металла_input.text() or 0)
            доработка_вырыв = int(self.доработка_вырыв_input.text() or 0)
            доработка_облой = int(self.доработка_облой_input.text() or 0)
            доработка_песок_на_поверхности = int(self.доработка_песок_на_поверхности_input.text() or 0)
            доработка_песок_в_резьбе = int(self.доработка_песок_в_резьбе_input.text() or 0)
            доработка_клей = int(self.доработка_клей_input.text() or 0)
            доработка_коробление = int(self.доработка_коробление_input.text() or 0)
            доработка_дефект_пеномодели = int(self.доработка_дефект_пеномодели_input.text() or 0)
            доработка_лапы = int(self.доработка_лапы_input.text() or 0)
            доработка_питатель = int(self.доработка_питатель_input.text() or 0)
            доработка_корона = int(self.доработка_корона_input.text() or 0)
            доработка_смещение = int(self.доработка_смещение_input.text() or 0)
            
            # Добавление окончательных браков в расчет
            окончательный_брак_недолив = int(self.окончательный_брак_недолив_input.text() or 0)
            окончательный_брак_вырыв = int(self.окончательный_брак_вырыв_input.text() or 0)
            окончательный_брак_зарез = int(self.окончательный_брак_зарез_input.text() or 0)
            окончательный_брак_коробление = int(self.окончательный_брак_коробление_input.text() or 0)
            окончательный_брак_наплыв_металла = int(self.окончательный_брак_наплыв_металла_input.text() or 0)
            окончательный_брак_нарушение_геометрии = int(self.окончательный_брак_нарушение_геометрии_input.text() or 0)
            окончательный_брак_нарушение_маркировки = int(self.окончательный_брак_нарушение_маркировки_input.text() or 0)
            окончательный_брак_непроклей = int(self.окончательный_брак_непроклей_input.text() or 0)
            окончательный_брак_неслитина = int(self.окончательный_брак_неслитина_input.text() or 0)
            окончательный_брак_несоответствие_внешнего_вида = int(self.окончательный_брак_несоответствие_внешнего_вида_input.text() or 0)
            окончательный_брак_несоответствие_размеров = int(self.окончательный_брак_несоответствие_размеров_input.text() or 0)
            окончательный_брак_пеномодель = int(self.окончательный_брак_пеномодель_input.text() or 0)
            окончательный_брак_пористость = int(self.окончательный_брак_пористость_input.text() or 0)
            окончательный_брак_пригар_песка = int(self.окончательный_брак_пригар_песка_input.text() or 0)
            окончательный_брак_прочее = int(self.окончательный_брак_прочее_input.text() or 0)
            окончательный_брак_рыхлота = int(self.окончательный_брак_рыхлота_input.text() or 0)
            окончательный_брак_раковины = int(self.окончательный_брак_раковины_input.text() or 0)
            окончательный_брак_скол = int(self.окончательный_брак_скол_input.text() or 0)
            окончательный_брак_слом = int(self.окончательный_брак_слом_input.text() or 0)
            окончательный_брак_спай = int(self.окончательный_брак_спай_input.text() or 0)
            окончательный_брак_трещины = int(self.окончательный_брак_трещины_input.text() or 0)

            # Расчет контроль_принято
            контроль_принято = контроль_отлито - (
                второй_сорт_раковины + второй_сорт_зарез +
                доработка_раковины + доработка_зарез +
                доработка_несоответствие_размеров + доработка_несоответствие_внешнего_вида +
                доработка_наплыв_металла + доработка_прорыв_металла +
                доработка_вырыв + доработка_облой +
                доработка_песок_на_поверхности + доработка_песок_в_резьбе +
                доработка_клей + доработка_коробление +
                доработка_дефект_пеномодели + доработка_лапы +
                доработка_питатель + доработка_корона +
                доработка_смещение + окончательный_брак_недолив + окончательный_брак_вырыв +
                окончательный_брак_зарез + окончательный_брак_коробление +
                окончательный_брак_наплыв_металла + окончательный_брак_нарушение_геометрии +
                окончательный_брак_нарушение_маркировки + окончательный_брак_непроклей +
                окончательный_брак_неслитина + окончательный_брак_несоответствие_внешнего_вида +
                окончательный_брак_несоответствие_размеров + окончательный_брак_пеномодель +
                окончательный_брак_пористость + окончательный_брак_пригар_песка +
                окончательный_брак_прочее + окончательный_брак_рыхлота +
                окончательный_брак_раковины + окончательный_брак_скол +
                окончательный_брак_слом + окончательный_брак_спай +
                окончательный_брак_трещины
            )
            self.контроль_принято_input.setText(str(контроль_принято))
        except ValueError:
            self.контроль_принято_input.setText("")

    def save_data(self):
        try:
            # Проверка обязательных полей
            if not self.номер_плавки_input.currentText():
                QMessageBox.warning(self, "Ошибка", "Выберите номер плавки")
                return
                
            if not self.контроль_отлито_input.text():
                QMessageBox.warning(self, "Ошибка", "Укажите количество отлитых деталей")
                return
                
            if not self.контролер1_input.currentText() and not self.контролер2_input.currentText() and not self.контролер3_input.currentText():
                QMessageBox.warning(self, "Ошибка", "Укажите хотя бы одного контролера")
                return

            # Собираем все данные в список
            data = [
                self.номер_плавки_input.currentText(),
                self.контроль_отлито_input.text(),
                self.контроль_принято_input.text(),
                self.контроль_дата_приемки_input.date().toString("dd.MM.yyyy"),
                self.контролер1_input.currentText(),
                self.контролер2_input.currentText(),
                self.контролер3_input.currentText(),
                self.второй_сорт_раковины_input.text(),
                self.второй_сорт_зарез_input.text(),
                self.доработка_раковины_input.text(),
                self.доработка_зарез_input.text(),
                self.доработка_несоответствие_размеров_input.text(),
                self.доработка_несоответствие_внешнего_вида_input.text(),
                self.доработка_наплыв_металла_input.text(),
                self.доработка_прорыв_металла_input.text(),
                self.доработка_вырыв_input.text(),
                self.доработка_облой_input.text(),
                self.доработка_песок_на_поверхности_input.text(),
                self.доработка_песок_в_резьбе_input.text(),
                self.доработка_клей_input.text(),
                self.доработка_коробление_input.text(),
                self.доработка_дефект_пеномодели_input.text(),
                self.доработка_лапы_input.text(),
                self.доработка_питатель_input.text(),
                self.доработка_корона_input.text(),
                self.доработка_смещение_input.text(),
                self.окончательный_брак_недолив_input.text(),
                self.окончательный_брак_вырыв_input.text(),
                self.окончательный_брак_зарез_input.text(),
                self.окончательный_брак_коробление_input.text(),
                self.окончательный_брак_наплыв_металла_input.text(),
                self.окончательный_брак_нарушение_геометрии_input.text(),
                self.окончательный_брак_нарушение_маркировки_input.text(),
                self.окончательный_брак_непроклей_input.text(),
                self.окончательный_брак_неслитина_input.text(),
                self.окончательный_брак_несоответствие_внешнего_вида_input.text(),
                self.окончательный_брак_несоответствие_размеров_input.text(),
                self.окончательный_брак_пеномодель_input.text(),
                self.окончательный_брак_пористость_input.text(),
                self.окончательный_брак_пригар_песка_input.text(),
                self.окончательный_брак_прочее_input.text(),
                self.окончательный_брак_рыхлота_input.text(),
                self.окончательный_брак_раковины_input.text(),
                self.окончательный_брак_скол_input.text(),
                self.окончательный_брак_слом_input.text(),
                self.окончательный_брак_спай_input.text(),
                self.окончательный_брак_трещины_input.text()
            ]

            headers = [
                'Номер_плавки', 'Контроль_отлито', 'Контроль_принято',
                'Контроль_дата_приемки', 'Контролер1', 'Контролер2', 'Контролер3',
                'Второй_сорт_раковины', 'Второй_сорт_зарез',
                'Доработка_раковины', 'Доработка_зарез',
                'Доработка_несоответствие_размеров', 'Доработка_несоответствие_внешнего_вида',
                'Доработка_наплыв_металла', 'Доработка_прорыв_металла',
                'Доработка_вырыв', 'Доработка_облой',
                'Доработка_песок_на_поверхности', 'Доработка_песок_в_резьбе',
                'Доработка_клей', 'Доработка_коробление',
                'Доработка_дефект_пеномодели', 'Доработка_лапы',
                'Доработка_питатель', 'Доработка_корона',
                'Доработка_смещение',
                'Окончательный_брак_недолив', 'Окончательный_брак_вырыв',
                'Окончательный_брак_зарез', 'Окончательный_брак_коробление',
                'Окончательный_брак_наплыв_металла', 'Окончательный_брак_нарушение_геометрии',
                'Окончательный_брак_нарушение_маркировки', 'Окончательный_брак_непроклей',
                'Окончательный_брак_неслитина', 'Окончательный_брак_несоответствие_внешнего_вида',
                'Окончательный_брак_несоответствие_размеров', 'Окончательный_брак_пеномодель',
                'Окончательный_брак_пористость', 'Окончательный_брак_пригар_песка',
                'Окончательный_брак_прочее', 'Окончательный_брак_рыхлота',
                'Окончательный_брак_раковины', 'Окончательный_брак_скол',
                'Окончательный_брак_слом', 'Окончательный_брак_спай',
                'Окончательный_брак_трещины'
            ]

            if os.path.exists('control.xlsx'):
                wb = load_workbook('control.xlsx')
                ws = wb.active  # Используем Sheet1
            else:
                wb = Workbook()
                ws = wb.active
                # Добавляем заголовки только если это новый файл
                for col, header in enumerate(headers, start=1):
                    ws.cell(row=1, column=col, value=header)

            # Добавляем новую строку данных
            next_row = ws.max_row + 1
            for col, value in enumerate(data, start=1):
                cell = ws.cell(row=next_row, column=col)
                cell.value = value
                if col == 4:  # Колонка D (дата)
                    cell.number_format = 'DD.MM.YYYY'

            # Применяем формат даты ко всем ячейкам в колонке D
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=4)
                cell.number_format = 'DD.MM.YYYY'

            wb.save('control.xlsx')
            wb.close()

            QMessageBox.information(self, "Успех", "Данные успешно сохранены!")
            
            # Очищаем и обновляем список доступных номеров плавок
            self.номер_плавки_input.clear()
            self.load_plavka_numbers()  # Обновляем список доступных номеров
            
            # Очищаем форму
            self.clear_form()
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при сохранении данных: {str(e)}")
        
        
    def clear_form(self):
        """Очистка формы"""
        # Очищаем все поля без дополнительных диалогов
        self.номер_плавки_input.setCurrentIndex(-1)
        self.контроль_отлито_input.setText('')
        self.контроль_принято_input.setText('')
        self.контроль_дата_приемки_input.setDate(QDate.currentDate())
        self.контролер1_input.setCurrentIndex(-1)
        self.контролер2_input.setCurrentIndex(-1)
        self.контролер3_input.setCurrentIndex(-1)

        # Очистка полей второго сорта
        self.второй_сорт_раковины_input.setText('')
        self.второй_сорт_зарез_input.setText('')

        # Очистка полей доработки
        self.доработка_раковины_input.setText('')
        self.доработка_зарез_input.setText('')
        self.доработка_несоответствие_размеров_input.setText('')
        self.доработка_несоответствие_внешнего_вида_input.setText('')
        self.доработка_наплыв_металла_input.setText('')
        self.доработка_прорыв_металла_input.setText('')
        self.доработка_вырыв_input.setText('')
        self.доработка_облой_input.setText('')
        self.доработка_песок_на_поверхности_input.setText('')
        self.доработка_песок_в_резьбе_input.setText('')
        self.доработка_клей_input.setText('')
        self.доработка_коробление_input.setText('')
        self.доработка_дефект_пеномодели_input.setText('')
        self.доработка_лапы_input.setText('')
        self.доработка_питатель_input.setText('')
        self.доработка_корона_input.setText('')
        self.доработка_смещение_input.setText('')

        # Очистка полей окончательного брака
        self.окончательный_брак_недолив_input.setText('')
        self.окончательный_брак_вырыв_input.setText('')
        self.окончательный_брак_зарез_input.setText('')
        self.окончательный_брак_коробление_input.setText('')
        self.окончательный_брак_наплыв_металла_input.setText('')
        self.окончательный_брак_нарушение_геометрии_input.setText('')
        self.окончательный_брак_нарушение_маркировки_input.setText('')
        self.окончательный_брак_непроклей_input.setText('')
        self.окончательный_брак_неслитина_input.setText('')
        self.окончательный_брак_несоответствие_внешнего_вида_input.setText('')
        self.окончательный_брак_несоответствие_размеров_input.setText('')
        self.окончательный_брак_пеномодель_input.setText('')
        self.окончательный_брак_пористость_input.setText('')
        self.окончательный_брак_пригар_песка_input.setText('')
        self.окончательный_брак_прочее_input.setText('')
        self.окончательный_брак_рыхлота_input.setText('')
        self.окончательный_брак_раковины_input.setText('')
        self.окончательный_брак_скол_input.setText('')
        self.окончательный_брак_слом_input.setText('')
        self.окончательный_брак_спай_input.setText('')
        self.окончательный_брак_трещины_input.setText('')

        # Обновляем наименование отливки
        self.update_наименование_отливки()

    def animate_group_hover(self, group, hover_in):
        if not hasattr(self, 'animations'):
            self.animations = []
        
        shadow = group.graphicsEffect()
        if shadow and shadow.isEnabled():
            # Удаляем завершенные анимации
            self.animations = [a for a in self.animations if a.state() != QPropertyAnimation.Stopped]
            
            animation = QPropertyAnimation(shadow, b"blurRadius")
            animation.setDuration(200)
            animation.setEasingCurve(QEasingCurve.InOutCubic)
            
            if hover_in:
                animation.setStartValue(15)
                animation.setEndValue(25)
            else:
                animation.setStartValue(25)
                animation.setEndValue(15)
            
            # Сохраняем анимацию
            self.animations.append(animation)
            animation.start()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    form = ControlForm()
    form.show()
    sys.exit(app.exec())